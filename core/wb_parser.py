"""
WB Parser — основной модуль парсинга Wildberries.
Использует Playwright для управления браузером.
"""
import threading
import time
import requests
import json
import os
import math
from datetime import datetime
import pandas as pd
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import urllib.request
from openpyxl.drawing.image import Image as OpenpyxlImage
import io
try:
    from PIL import Image as PILImage
except ImportError:
    PILImage = None


class WBParser:
    """Парсер товаров Wildberries."""

    CATEGORIES_URL = "https://static-basket-01.wbbasket.ru/vol0/data/main-menu-ru-ru-v3.json"
    RESULT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "result")

    def __init__(self, api):
        self.api = api
        self.is_running = False
        self.thread = None
        self.categories_cache = []
        self._browser = None
        self._playwright = None

    def fetch_categories(self):
        """Загружает дерево категорий с WB API. Возвращает кэш при повторных вызовах."""
        if self.categories_cache:
            return self.categories_cache

        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            )
        }

        urls = [
            "https://static-basket-01.wbbasket.ru/vol0/data/main-menu-ru-ru-v3.json",
            "https://static-basket-02.wbbasket.ru/vol0/data/main-menu-ru-ru-v3.json",
        ]

        data = None
        for url in urls:
            try:
                resp = requests.get(url, headers=headers, timeout=15)
                resp.raise_for_status()
                data = resp.json()
                break
            except Exception:
                continue

        if data is None:
            self.api.log("Ошибка загрузки категорий: нет ответа от сервера WB.")
            return []

        def process_nodes(nodes):
            """Рекурсивно обрабатывает узлы дерева категорий."""
            result = []
            for node in nodes:
                name = node.get("name", "")
                name_lower = name.lower()
                if not name or "сертификат" in name_lower or "экспресс" in name_lower or "ресейл" in name_lower:
                    continue
                item = {
                    "id": node.get("id", hash(name)),
                    "name": name,
                    "url": node.get("url", ""),
                    "shard": node.get("shard", ""),
                    "query": node.get("query", ""),
                }
                children = node.get("childs", [])
                if children:
                    item["children"] = process_nodes(children)
                result.append(item)
            return result

        self.categories_cache = process_nodes(data)
        return self.categories_cache

    def start(self, params):
        """Запускает парсинг в фоновом потоке."""
        if self.is_running:
            return
        self.is_running = True
        self.thread = threading.Thread(target=self._run, args=(params,), daemon=True)
        self.thread.start()

    def stop(self):
        """Останавливает парсинг. Принудительно закрывает браузер."""
        self.is_running = False
        self.api.log("Остановка парсинга...")
        try:
            if self._browser:
                self._browser.close()
                self._browser = None
        except Exception:
            pass

    def _find_category_by_id(self, nodes, target_id):
        """Ищет категорию по ID в дереве."""
        for node in nodes:
            if node["id"] == target_id:
                return node
            if "children" in node:
                res = self._find_category_by_id(node["children"], target_id)
                if res:
                    return res
        return None

    def _extract_digits(self, text: str) -> int | None:
        """Извлекает число из строки с ценой. Возвращает None если не найдено."""
        if not text:
            return None
        digits = "".join(filter(str.isdigit, text))
        return int(digits) if digits else None

    def _clean_name(self, raw: str) -> str:
        """Удаляет ведущий слеш и пробелы из названия товара."""
        name = raw.strip()
        if name.startswith("/ "):
            name = name[2:]
        elif name.startswith("/"):
            name = name[1:]
        return name.strip()

    def _ensure_playwright_browsers(self):
        """Проверяет и при необходимости автоматически устанавливает браузер Chromium для Playwright."""
        try:
            import os
            import platform

            if platform.system() == 'Windows':
                app_data = os.getenv('LOCALAPPDATA', os.path.expanduser('~'))
            else:
                app_data = os.path.expanduser('~')
            
            browsers_path = os.path.join(app_data, 'WB_Parser_Browsers')
            os.environ['PLAYWRIGHT_BROWSERS_PATH'] = browsers_path

            from playwright._impl._driver import compute_driver_executable
            import subprocess
            driver_executable, driver_cli = compute_driver_executable()
            
            with sync_playwright() as pw:
                try:
                    browser = pw.chromium.launch(headless=True)
                    browser.close()
                    return  
                except Exception as e:
                    if "Executable doesn't exist" in str(e) or "not installed" in str(e).lower():
                        self.api.log("Установка веб-драйвера Chromium")
                        
                        kwargs = {
                            "stdout": subprocess.DEVNULL,
                            "stderr": subprocess.DEVNULL,
                            "check": True
                        }
                        if platform.system() == 'Windows':
                            kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW

                        subprocess.run(
                            [driver_executable, driver_cli, "install", "chromium"],
                            **kwargs
                        )
                        self.api.log("Установка веб-драйвера завершена!")
        except Exception as e:
            self.api.log(f"Предупреждение при автоустановке браузера: {e}")

    def _run(self, params):
        """Основной метод парсинга. Запускается в отдельном потоке."""
        start_time = time.time()
        self.api.log("Запуск парсера...")
        all_results = []
        try:
            self._ensure_playwright_browsers()
            with sync_playwright() as pw:
                self._playwright = pw
                self._browser = pw.chromium.launch(headless=False)

                mode = params.get("mode", "all")
                price_min = params.get("price_min", 0)
                price_max = params.get("price_max", 100000)
                target_count = params.get("items_count", 100)
                selected_ids = params.get("selected_categories", [])

                if mode == "seller":
                    seller_url = params.get("seller_url", "").strip()
                    if not seller_url:
                        self.api.log("Ошибка: не указана ссылка на магазин/бренд.")
                        return

                    context = self._browser.new_context(viewport={"width": 1920, "height": 1080})
                    page = context.new_page()

                    target_url = seller_url
                    if "brands/" in seller_url:
                        brand_slug = seller_url.rstrip("/").split("brands/")[-1]
                        target_url = f"https://www.wildberries.ru/catalog/0/search.aspx?search={brand_slug}"
                        self.api.log(f"Парсинг каталога бренда «{brand_slug}»: {target_url}")
                    elif not seller_url.startswith("http"):
                        target_url = f"https://www.wildberries.ru/catalog/0/search.aspx?search={seller_url}"
                        self.api.log(f"Парсинг по запросу: {target_url}")
                    else:
                        self.api.log(f"Парсинг магазина: {seller_url}")

                    self._parse_target(page, target_url, target_count, price_min, price_max, "Магазин", all_results)

                else: 
                    context = self._browser.new_context(viewport={"width": 1920, "height": 1080})
                    page = context.new_page()

                    if not selected_ids:
                        self.api.log("Ошибка: не выбраны категории.")
                        return

                    for cat_id in selected_ids:
                        if not self.is_running:
                            break
                        cat_info = self._find_category_by_id(self.categories_cache, cat_id)
                        if cat_info and "children" not in cat_info:
                            url = f"https://www.wildberries.ru{cat_info['url']}"
                            self.api.log(f"Начинаем парсинг категории: {cat_info['name']}")
                            self._parse_target(page, url, target_count, price_min, price_max, cat_info["name"], all_results)

                if self._browser:
                    self._browser.close()
                    self._browser = None

        except Exception as e:
            if self.is_running: 
                self.api.log(f"Критическая ошибка: {e}")
        finally:
            if all_results:
                self._export_to_excel(all_results, params)
            elif self.is_running:
                self.api.log("Товары не найдены.")

            self.is_running = False
            self._browser = None
            self._playwright = None
            elapsed = int(time.time() - start_time)
            minutes, seconds = divmod(elapsed, 60)
            self.api.log(f"Работа завершена. Затрачено времени: {minutes} мин {seconds} сек.")
            self.api.notify_finished(getattr(self, "_last_exported_file", None))


    def _parse_target(self, page, url, target_count, price_min, price_max, cat_name, all_results):
        """
        Парсит страницу (категория или магазин) и добавляет товары в all_results.
        Поддерживает бесконечный скролл и переход по страницам.
        """
        try:
            self.api.log(f"Открытие страницы: {url}")
            page.goto(url, timeout=45000, wait_until="domcontentloaded")

            loaded = False
            for attempt in range(3):
                if not self.is_running:
                    return
                try:
                    page.wait_for_selector(".product-card", timeout=12000)
                    loaded = True
                    break
                except PlaywrightTimeout:
                    if attempt < 2:
                        self.api.log(f"Ожидание товаров... попытка {attempt + 2}/3")
                        time.sleep(2)

            if not loaded:
                self.api.log(f"Товары на странице {url} не найдены (антибот или пустая категория).")
                return

            collected = 0
            processed_urls: set[str] = set()
            no_new_scrolls = 0
            current_page_num = 1

            while collected < target_count and self.is_running:
                try:
                    page.wait_for_selector(".product-card", timeout=8000)
                except PlaywrightTimeout:
                    break

                cards = page.locator(".product-card").all()
                new_found = False

                for card in cards:
                    if collected >= target_count or not self.is_running:
                        break

                    try:
                        link_el = card.locator("a.product-card__link, a.j-card-link")
                        item_url = link_el.get_attribute("href")
                        if not item_url:
                            continue
                        if item_url.startswith("/"):
                            item_url = "https://www.wildberries.ru" + item_url

                        if item_url in processed_urls:
                            continue
                        processed_urls.add(item_url)
                        new_found = True

                        wallet_text_el = card.locator(".price__wallet-condition-text, .price__wallet-condition-text.red-price")
                        has_wallet = wallet_text_el.count() > 0 and "Кошельком" in wallet_text_el.first.text_content()

                        if has_wallet:
                            wallet_el = card.locator("ins.wallet-price")
                            wallet_raw = self._extract_digits(
                                wallet_el.text_content() if wallet_el.count() > 0 else ""
                            )
                            lower_el = card.locator(".price__lower-price")
                            sale_raw = self._extract_digits(
                                lower_el.text_content() if lower_el.count() > 0 else ""
                            )
                            if wallet_raw and sale_raw and sale_raw != wallet_raw:
                                wallet_price = wallet_raw
                                sale_price = sale_raw
                            elif wallet_raw:
                                wallet_price = wallet_raw
                                sale_price = wallet_raw
                            else:
                                continue
                        else:
                            lower_el = card.locator("ins.wallet-price, .price__lower-price")
                            sale_raw = self._extract_digits(
                                lower_el.first.text_content() if lower_el.count() > 0 else ""
                            )
                            if not sale_raw:
                                continue
                            wallet_price = "-"
                            sale_price = sale_raw


                        lower_price = sale_price  

                        del_el = card.locator("del")
                        original_price = self._extract_digits(
                            del_el.first.text_content() if del_el.count() > 0 else ""
                        )
                        if not original_price:
                            original_price = sale_price

                        if not (price_min <= sale_price <= price_max):
                            continue

                        name_el = card.locator(".product-card__name")
                        name = self._clean_name(
                            name_el.text_content() if name_el.count() > 0 else "Без названия"
                        )

                        img_url = ""
                        img_el = card.locator("img.j-thumbnail, .product-card__img img")
                        if img_el.count() > 0:
                            src = img_el.first.get_attribute("src") or img_el.first.get_attribute("data-src-pb")
                            if src:
                                img_url = ("https:" + src) if src.startswith("//") else src

                        all_results.append({
                            "Изображение": img_url,
                            "Категория": cat_name,
                            "Название товара": name,
                            "Цена со скидкой": sale_price,
                            "Цена без скидки": original_price,
                            "Цена с WB кошельком": wallet_price,
                            "Ссылка": item_url,
                        })
                        collected += 1
                        if collected % 10 == 0:
                            self.api.log(f"Собрано {collected} товаров из «{cat_name}»...")

                    except Exception:
                        continue  

                if not new_found:
                    no_new_scrolls += 1
                    if no_new_scrolls > 4:
                        next_btn = page.locator("a.pagination-next, button.pagination-next")
                        if next_btn.count() > 0 and next_btn.is_enabled():
                            current_page_num += 1
                            self.api.log(f"Переход на страницу {current_page_num}...")
                            next_btn.click()
                            time.sleep(3)
                            no_new_scrolls = 0
                        else:
                            self.api.log(f"Больше товаров не найдено в «{cat_name}».")
                            break
                else:
                    no_new_scrolls = 0
                    page.evaluate("window.scrollBy(0, window.innerHeight)")
                    time.sleep(1.5)

        except Exception as e:
            if self.is_running:
                self.api.log(f"Ошибка при парсинге {url}: {e}")

    def _export_to_excel(self, data: list, params: dict):
        """Формирует и сохраняет Excel-файл с результатами парсинга."""
        try:
            self.api.log("Формирование Excel файла...")
            df = pd.DataFrame(data)
            df = df.sort_values(by="Цена со скидкой")

            columns_order = [
                "Изображение",
                "Название товара",
                "Цена без скидки",
                "Цена со скидкой",
                "Цена с WB кошельком",
                "Ссылка",
            ]
            columns_order = [c for c in columns_order if c in df.columns]

            os.makedirs(self.RESULT_DIR, exist_ok=True)
            timestamp = datetime.now().strftime("%d.%m.%Y_%H-%M")

            mode = params.get("mode", "all")
            if mode == "seller":
                raw_name = params.get("seller_url", "Магазин").rstrip("/").split("/")[-1]
                safe = "".join(c for c in raw_name if c.isalnum() or c in "-_")[:50] or "Магазин"
                filepath = os.path.join(self.RESULT_DIR, f"{safe}_{timestamp}.xlsx")
                self._save_xlsx(df[columns_order], filepath, params)
            elif params.get("merge_files", True):
                selected_cat_ids = params.get("selected_categories", [])
                cat_names = []
                for cid in selected_cat_ids:
                    info = self._find_category_by_id(self.categories_cache, cid)
                    if info and "children" not in info:
                        safe_n = "".join(c for c in info["name"] if c.isalnum() or c == " ").strip()
                        cat_names.append(safe_n)
                joined = "_".join(cat_names)[:60] or "Все"
                filepath = os.path.join(self.RESULT_DIR, f"{joined}_{timestamp}.xlsx")
                filepath = self._save_xlsx(df[columns_order], filepath, params)
            else:
                filepath = ""
                for cat_name, group in df.groupby("Категория"):
                    safe = "".join(c for c in str(cat_name) if c.isalnum() or c == " ").strip()[:50]
                    filepath = os.path.join(self.RESULT_DIR, f"{safe}_{timestamp}.xlsx")
                    filepath = self._save_xlsx(group[columns_order], filepath, params)

            self._last_exported_file = filepath
            if filepath:
                self.api.log(f"Excel сохранён: {filepath}")

        except Exception as e:
            self.api.log(f"Ошибка при формировании Excel: {e}")

    def _save_xlsx(self, df: pd.DataFrame, filepath: str, params: dict) -> str:
        """
        Сохраняет DataFrame в Excel и вставляет превью изображений товаров.
        Если файл открыт в Microsoft Excel, сохраняет с суффиксом.
        """
        save_images = params.get("save_images", True)
        
        if not save_images and "Изображение" in df.columns:
            df = df.drop(columns=["Изображение"])
            
        try:
            df.to_excel(filepath, index=False)
        except PermissionError:
            time_suffix = datetime.now().strftime("_%S")
            base, ext = os.path.splitext(filepath)
            filepath = f"{base}{time_suffix}{ext}"
            self.api.log("Предупреждение: файл был открыт в Excel. Сохраняем в новый файл!")
            df.to_excel(filepath, index=False)

        try:
            wb = load_workbook(filepath)
            ws = wb.active

            wrap_alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            left_alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
            
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.alignment = wrap_alignment

            for col_idx, col_name in enumerate(df.columns, start=1):
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                if col_name == "Изображение":
                    ws.column_dimensions[col_letter].width = 14
                elif col_name == "Название товара":
                    ws.column_dimensions[col_letter].width = 45
                elif col_name == "Ссылка":
                    ws.column_dimensions[col_letter].width = 30
                else: # Цены
                    ws.column_dimensions[col_letter].width = 18

                for row_idx in range(2, len(df) + 2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if col_name == "Название товара":
                        cell.alignment = left_alignment
                    else:
                        cell.alignment = wrap_alignment

            for row_idx in range(2, len(df) + 2):
                ws.row_dimensions[row_idx].height = 72

                if not save_images or PILImage is None:
                    continue

                if "Изображение" not in df.columns:
                    continue
                    
                img_col_idx = df.columns.get_loc("Изображение") + 1
                img_col_letter = ws.cell(row=1, column=img_col_idx).column_letter
                
                img_val = df.iloc[row_idx - 2].get("Изображение", "")
                if not (img_val and isinstance(img_val, str) and img_val.startswith("http")):
                    continue

                try:
                    req = urllib.request.Request(img_val, headers={"User-Agent": "Mozilla/5.0"})
                    with urllib.request.urlopen(req, timeout=6) as resp:
                        raw = io.BytesIO(resp.read())
                    pil_img = PILImage.open(raw)
                    
                    if pil_img.mode in ('RGBA', 'P', 'LA') or pil_img.format == 'WEBP':
                        pil_img = pil_img.convert('RGB')
                        
                    pil_img.thumbnail((90, 90))
                    out = io.BytesIO()
                    pil_img.save(out, format="JPEG")
                    out.seek(0)
                    xl_img = OpenpyxlImage(out)
                    ws.add_image(xl_img, f"{img_col_letter}{row_idx}")
                    ws[f"{img_col_letter}{row_idx}"].value = ""  
                except Exception as e:
                    pass  

            wb.save(filepath)
        except Exception as e:
            print(f"[XLSX] Ошибка при вставке изображений: {e}")

        return filepath
